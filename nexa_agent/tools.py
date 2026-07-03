"""
ReAct Agent 工具集

提供 web_search (Tavily + DuckDuckGo), wikipedia_search (REST API),
analyze_image (端侧 MiniCPM-V), analyze_image_cloud (云端 Kimi K2.6),
tavily_extract + save_content (网页提取+落盘), calculator, get_current_time,
read_pdf (pymupdf4llm), read_xlsx (openpyxl/csv) 工具的统一注册与执行。

每个工具函数签名统一为:
    def tool_name(param: str) -> str

返回值均为 Observation 字符串（成功或错误消息）。
"""

from __future__ import annotations

import base64
import os
import re
import time
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional

import requests

# 加载 .env（兼容独立调用场景）
try:
    from dotenv import load_dotenv

    _project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _dotenv_path = os.path.join(_project_root, ".env")
    if os.path.exists(_dotenv_path):
        load_dotenv(_dotenv_path)
except ImportError:
    pass

from nexa_agent.logger import get_logger

logger = get_logger("react_tools")

# ==========================================================================
# 工具注册表
# ==========================================================================

# 全局工具注册表: tool_name -> callable
TOOLS: Dict[str, Callable[[str], str]] = {}

# 工具元数据: tool_name -> {description, signature, examples}
TOOL_META: Dict[str, Dict[str, Any]] = {}


def register(name: str, description: str, signature: str, examples: List[str]):
    """装饰器：将函数注册到 TOOLS 全局表中"""

    def decorator(func: Callable[[str], str]):
        TOOLS[name] = func
        TOOL_META[name] = {
            "description": description,
            "signature": signature,
            "examples": examples,
        }
        return func

    return decorator


def get_openai_tool_definitions() -> List[Dict[str, Any]]:
    """生成 OpenAI function calling 格式的工具定义列表"""
    definitions = []
    for name, meta in TOOL_META.items():
        desc = meta["description"]
        if meta.get("examples"):
            desc += "\n示例: " + "; ".join(meta["examples"])

        # 所有工具都接受单个字符串参数（保持与现有接口兼容）
        tool_def = {
            "type": "function",
            "function": {
                "name": name,
                "description": desc,
                "parameters": {
                    "type": "object",
                    "properties": {
                        "input": {
                            "type": "string",
                            "description": _get_param_description(name, meta),
                        }
                    },
                    "required": ["input"] if name != "get_current_time" else [],
                },
            },
        }
        definitions.append(tool_def)

    # 显式终止工具：结构化提交裁定并结束调查（由 react_loop 特殊处理，不进 TOOLS 执行）
    definitions.append(_submit_verdict_tool_def())
    return definitions


def _submit_verdict_tool_def() -> Dict[str, Any]:
    """submit_verdict 结构化终止工具定义。

    模型在充分取证后调用它提交最终裁定，循环即结束。相比自由文本终止，
    它给出结构化 schema（便于解析/校验），并作为强制取证 gate 与来源对账的落点。
    """
    return {
        "type": "function",
        "function": {
            "name": "submit_verdict",
            "description": (
                "提交最终裁定并结束本次调查。**仅在你已通过检索工具充分取证后调用。**"
                "每条 evidence 必须绑定你真实调用过的工具/URL 作为来源，不得臆造。"
                "证据不足时用 need_user_confirm 如实列出，不要猜测。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "verdict": {
                        "type": "string",
                        "description": "总体裁定：靠谱 / 存疑 / 大概率有坑（或任务要求的结论标签）",
                    },
                    "summary": {"type": "string", "description": "一句话结论与核心理由"},
                    "evidence": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "证据条目，每条含：事实 + 真实来源（调用过的工具/URL）+ 置信度",
                    },
                    "red_flags": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "发现的红旗信号（可空）",
                    },
                    "need_user_confirm": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "证据不足、需用户自行确认的事项（可空）",
                    },
                },
                "required": ["verdict", "summary", "evidence"],
            },
        },
    }


def _get_param_description(name: str, meta: dict) -> str:
    """根据工具名生成参数描述"""
    sig = meta.get("signature", "")
    param_map = {
        "web_search": "搜索关键词",
        "wikipedia_search": "搜索关键词",
        "web_fetch": "要抓取的网页 URL",
        "tavily_extract": "要提取内容的网页 URL",
        "save_content": "保存文件名（不含扩展名）",
        "analyze_image": "图片路径或 URL | 分析提示词（用 | 分隔）",
        "analyze_image_cloud": "图片路径或 URL | 分析提示词（用 | 分隔）",
        "calculator": "数学表达式，支持 sqrt/sin/cos/log/pi 等",
        "get_current_time": "无需参数",
        "domain_whois_lookup": "域名、URL 或邮箱地址（自动提取其中的注册域名）",
        "read_pdf": "PDF 路径或 URL。大文档可指定页码: path | pages=3-5",
        "read_xlsx": "Excel 文件路径或 URL",
    }
    return param_map.get(name, f"参数，格式参考: {sig}")


# ==========================================================================
# 工具实现
# ==========================================================================


@register(
    name="web_search",
    description="搜索互联网，返回前若干条实时结果（含内容摘要）；内部经可插拔 provider 层多源降级（Tavily→自建 SearXNG→Exa→DuckDuckGo）",
    signature="web_search(query)",
    examples=["web_search(2025年诺贝尔奖得主)"],
)
def web_search(query: str) -> str:
    """Web 搜索 — 经可插拔 provider 层（Tavily→SearXNG→Exa→DDG 有序降级）。

    provider 顺序、SearXNG 地址、熔断阈值等由 config.SEARCH_CONFIG 控制。

    Args:
        query: 搜索关键词

    Returns:
        前若干条结果的标题、链接、摘要
    """
    query = query.strip()
    if not query:
        return "[错误] web_search: 查询内容不能为空"

    logger.info("web_search 调用 query=%s", query)

    from nexa_agent.config import SEARCH_CONFIG
    from nexa_agent.search import enrich_results, get_default_router

    try:
        router = get_default_router()
        results, provider = router.search(query, SEARCH_CONFIG["max_results"])
    except Exception as exc:
        logger.error("web_search 失败: %s", exc, exc_info=True)
        return f"[错误] web_search 执行失败: {exc}"

    if not results:
        return f"未找到与 '{query}' 相关的结果（所有搜索 provider 均无结果或不可用）。"

    # 增强层：对摘要偏弱的 provider（如自建 SearXNG / DDG）抓取正文补齐
    if SEARCH_CONFIG["enrich_enabled"] and provider in SEARCH_CONFIG["enrich_providers"]:
        results = enrich_results(
            results,
            top_k=SEARCH_CONFIG["enrich_top_k"],
            max_chars=SEARCH_CONFIG["enrich_max_chars"],
            timeout=SEARCH_CONFIG["enrich_timeout"],
        )

    snippet_max = SEARCH_CONFIG["snippet_max_chars"]
    lines = [f"{provider} 搜索 '{query}' 结果（共 {len(results)} 条）:\n"]
    for i, r in enumerate(results, 1):
        snippet = r.snippet
        # 已增强的摘要保留全文（增强层已按 enrich_max_chars 截断）；其余截到 snippet_max
        if not r.enriched and len(snippet) > snippet_max:
            snippet = snippet[:snippet_max] + "..."
        lines.append(f"{i}. {r.title}")
        lines.append(f"   链接: {r.url}")
        if snippet:
            label = "正文摘要" if r.enriched else "摘要"
            lines.append(f"   {label}: {snippet}")
        lines.append("")
    logger.info("web_search 完成 provider=%s results=%d", provider, len(results))
    return "\n".join(lines)


@register(
    name="wikipedia_search",
    description="搜索 Wikipedia，返回最相关文章的摘要（最多 800 字符）",
    signature="wikipedia_search(query)",
    examples=["wikipedia_search(transformer deep learning)"],
)
def wikipedia_search(query: str) -> str:
    """Wikipedia 百科搜索

    直接调用 Wikipedia REST API，根据 query 自动选择中文/英文端点。
    只返回条目的导言纯文本，避免内容过长。

    Args:
        query: 搜索关键词（支持中英文）

    Returns:
        条目标题 + 纯文本摘要（~800 字符）
    """
    query = query.strip()
    if not query:
        return "[错误] wikipedia_search: 查询内容不能为空"

    logger.info("wikipedia_search 调用 query=%s", query)

    # 自动检测语言
    has_chinese = bool(re.search(r"[一-鿿]", query))
    lang = "zh" if has_chinese else "en"
    api_url = f"https://{lang}.wikipedia.org/w/api.php"

    headers = {"User-Agent": "NexaAgent-ReAct/1.0 (nexa@example.com)"}

    try:
        # 第一步：搜索最相关的条目
        search_res = requests.get(
            api_url,
            params={
                "action": "query",
                "list": "search",
                "srsearch": query,
                "format": "json",
                "srlimit": 1,
            },
            headers=headers,
            timeout=10,
        )
        search_res.raise_for_status()
        search_hits = search_res.json().get("query", {}).get("search", [])

        if not search_hits:
            # 中文没找到，回退英文
            if lang == "zh":
                logger.info("中文 Wikipedia 无结果，回退英文")
                api_url = "https://en.wikipedia.org/w/api.php"
                search_res = requests.get(
                    api_url,
                    params={
                        "action": "query", "list": "search",
                        "srsearch": query, "format": "json", "srlimit": 1,
                    },
                    headers=headers, timeout=10,
                )
                search_res.raise_for_status()
                search_hits = search_res.json().get("query", {}).get("search", [])
            if not search_hits:
                return f"Wikipedia 未找到与 '{query}' 相关的条目。"

        target_title = search_hits[0]["title"]

        # 第二步：获取条目导言（纯文本）
        content_res = requests.get(
            api_url,
            params={
                "action": "query",
                "prop": "extracts",
                "exintro": True,
                "explaintext": True,
                "titles": target_title,
                "format": "json",
            },
            headers=headers,
            timeout=10,
        )
        content_res.raise_for_status()
        pages = content_res.json().get("query", {}).get("pages", {})

        for page_id, page_info in pages.items():
            if page_id == "-1":
                return f"Wikipedia 条目 '{target_title}' 内容无法读取。"
            extract = (page_info.get("extract") or "").strip()
            if extract:
                # 截断到 800 字符
                if len(extract) > 800:
                    extract = extract[:800] + "..."
                return f"Wikipedia - {target_title}\n链接: https://{lang}.wikipedia.org/wiki/{requests.utils.quote(target_title)}\n\n{extract}"

        return f"Wikipedia 找到条目 '{target_title}'，但无正文摘要。"

    except requests.exceptions.Timeout:
        logger.error("Wikipedia API 超时")
        return "[错误] wikipedia_search: Wikipedia API 请求超时，请稍后重试"
    except requests.exceptions.ConnectionError:
        logger.error("Wikipedia API 连接失败（网络可能受限）")
        return "[错误] wikipedia_search: 无法连接 Wikipedia，网络可能受限，请尝试 web_search"
    except Exception as exc:
        logger.error("wikipedia_search 失败: %s", exc, exc_info=True)
        return f"[错误] wikipedia_search 执行失败: {exc}"


# --------------------------------------------------------------------------
# 图片工具辅助函数
# --------------------------------------------------------------------------

def _download_image_from_url(url: str) -> str:
    """下载 URL 图片到临时文件，返回本地路径。调用方负责清理。"""
    import tempfile
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    }
    resp = requests.get(url, timeout=60, headers=headers)
    if resp.status_code != 200:
        raise RuntimeError(f"图片下载失败 HTTP {resp.status_code} — {url[:100]}")
    if len(resp.content) < 100:
        raise RuntimeError(f"图片下载内容太小 ({len(resp.content)} bytes)")

    ext = ".png"
    content_type = resp.headers.get("Content-Type", "")
    for candidate_ext, mime in [(".jpg", "jpeg"), (".png", "png"), (".webp", "webp"), (".gif", "gif")]:
        if mime in content_type:
            ext = candidate_ext
            break
    else:
        url_lower = url.lower().split("?")[0]
        for candidate_ext in [".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"]:
            if url_lower.endswith(candidate_ext):
                ext = candidate_ext
                break

    tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
    tmp.write(resp.content)
    tmp.close()
    logger.info("_download_image_from_url 已下载 %d bytes → %s", len(resp.content), tmp.name)
    return tmp.name


def _resolve_image_path(raw_path: str) -> str:
    """解析图片路径为绝对路径

    支持四种形式:
      - URL: http(s)://... (自动下载到临时文件)
      - 绝对路径: /path/to/image.jpg
      - 相对项目根: ./data/image.jpg or data/image.jpg
      - 相对当前目录: image.jpg

    返回解析后的绝对路径，若文件不存在则抛出 FileNotFoundError。
    URL 下载的临时文件路径以 _IMAGE_TMP_PREFIX 标记，由调用方清理。
    """
    raw_path = raw_path.strip()

    if raw_path.startswith("http://") or raw_path.startswith("https://"):
        return _download_image_from_url(raw_path)

    # 去掉开头的 ./
    if raw_path.startswith("./"):
        raw_path = raw_path[2:]

    # 绝对路径直接返回
    if os.path.isabs(raw_path):
        if os.path.isfile(raw_path):
            return raw_path
        raise FileNotFoundError(f"图片文件不存在: {raw_path}")

    # 尝试相对项目根目录
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    candidates = [
        os.path.join(project_root, raw_path),          # 相对项目根
        os.path.join(project_root, "data", raw_path),  # data/ 子目录
        os.path.join(os.getcwd(), raw_path),            # 相对当前工作目录
    ]

    for p in candidates:
        if os.path.isfile(p):
            return os.path.abspath(p)

    raise FileNotFoundError(
        f"图片文件不存在，尝试过以下路径:\n"
        + "\n".join(f"  - {c}" for c in candidates)
    )


def _parse_image_tool_args(param: str) -> tuple:
    """解析图片工具的 | 分隔参数

    Args:
        param: 原始参数字符串，如 "path/to/img.jpg | 描述这张图"

    Returns:
        (image_path: str, prompt: str)
    """
    if "|" in param:
        parts = param.split("|", 1)
        image_path = parts[0].strip()
        prompt = parts[1].strip()
    else:
        image_path = param.strip()
        prompt = "请描述这张图片的内容"

    return image_path, prompt


# --------------------------------------------------------------------------
# 端侧图片分析
# --------------------------------------------------------------------------

@register(
    name="analyze_image",
    description="使用端侧本地 MiniCPM 视觉模型分析图片，速度快，适合简单描述或轻量感知。"
                "支持本地路径和 URL（自动下载）。",
    signature="analyze_image(image_path_or_url | prompt)",
    examples=[
        "analyze_image(./data/photo.jpg | 描述图中的主要内容)",
        "analyze_image(https://example.com/chart.png | 提取图中文字)",
    ],
)
def analyze_image(param: str) -> str:
    """端侧图片分析 — llama.cpp MiniCPM-V

    通过 llama.cpp server 的 OpenAI 兼容 API 调用本地 VLM。
    支持本地路径和 URL（自动下载）。
    """
    param = param.strip()
    if not param:
        return "[错误] analyze_image: 参数不能为空，格式为 '图片路径或URL | 提示词'"

    image_path, prompt = _parse_image_tool_args(param)

    _img_exts = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
    _check_ext = os.path.splitext(image_path.split("?")[0])[1].lower()
    if _check_ext and _check_ext not in _img_exts:
        return (
            f"[错误] analyze_image: 不支持的文件类型 '{_check_ext}'。"
            f"此工具仅接受图片文件 {sorted(_img_exts)}。"
            f"Excel 请用 read_xlsx，PDF 请用 read_pdf。"
        )

    logger.info("analyze_image 调用 path=%s prompt=%s", image_path, prompt[:100])

    is_url = image_path.startswith("http://") or image_path.startswith("https://")
    resolved_path = None

    try:
        resolved_path = _resolve_image_path(image_path)
    except (FileNotFoundError, RuntimeError) as e:
        logger.error("analyze_image: %s", e)
        return f"[错误] analyze_image: {e}"

    vlm_base_url = os.environ.get("VLM_BASE_URL", "http://127.0.0.1:8080/v1")
    vlm_model = os.environ.get("VLM_MODEL_NAME", "minicpm-v")
    vlm_ctx_size = int(os.environ.get("VLM_CTX_SIZE", "4096"))

    try:
        from openai import OpenAI

        client = OpenAI(
            api_key="not-needed",
            base_url=vlm_base_url,
            timeout=120.0,
        )

        with open(resolved_path, "rb") as f:
            image_data = base64.b64encode(f.read()).decode("utf-8")

        ext = resolved_path.rsplit(".", 1)[-1].lower() if "." in resolved_path else "png"
        mime_map = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "webp": "webp"}
        mime_type = f"image/{mime_map.get(ext, 'png')}"

        t0 = time.time()
        response = client.chat.completions.create(
            model=vlm_model,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:{mime_type};base64,{image_data}"},
                        },
                    ],
                }
            ],
            max_tokens=vlm_ctx_size,
            temperature=0.1,
            stream=False,
            extra_body={"n_ctx": vlm_ctx_size},
        )

        elapsed_ms = (time.time() - t0) * 1000
        content = response.choices[0].message.content or "(VLM 返回空内容)"
        tokens = response.usage.total_tokens if response.usage else 0

        logger.info(
            "analyze_image 完成 elapsed=%.0fms tokens=%d model=%s",
            elapsed_ms, tokens, vlm_model,
        )

        return content

    except ImportError:
        return "[错误] analyze_image: openai 库未安装"
    except Exception as exc:
        logger.error("analyze_image VLM 调用失败: %s", exc, exc_info=True)
        return f"[错误] analyze_image 执行失败: {exc}"
    finally:
        if is_url and resolved_path and os.path.exists(resolved_path):
            os.unlink(resolved_path)


# --------------------------------------------------------------------------
# 云端图片分析 — GMI Gemini 3.1（有 GMI key 时）/ Kimi K2.6（回落）
# --------------------------------------------------------------------------

@register(
    name="analyze_image_cloud",
    description="使用云端视觉大模型分析图片，适合 OCR、复杂场景理解。"
                "默认走 GMI Gemini 3.1（普通图用 flash-lite，空响应自动升级 pro），"
                "无 GMI key 时回落 Kimi 多模态。支持本地路径和 URL（自动下载）。",
    signature="analyze_image_cloud(image_path_or_url | prompt)",
    examples=[
        "analyze_image_cloud(./data/chart.png | 提取图中所有文字并解释含义)",
        "analyze_image_cloud(https://example.com/doc.jpg | OCR 提取所有文字)",
    ],
)
def analyze_image_cloud(param: str) -> str:
    """云端图片分析 — GMI Gemini 3.1（默认）/ Kimi K2.6（回落）

    按 config.VISION_CONFIG 选 provider：有 GMI_API_KEY 时走 GMI 的
    gemini-3.1-flash-lite-preview（普通图）；返回空内容时自动升级
    gemini-3.1-pro-preview（复杂/模糊图）重试一次。无 GMI key 回落 Kimi。

    Args:
        param: "图片路径 | 提示词" 格式的字符串

    Returns:
        视觉模型的分析文本
    """
    from nexa_agent.config import VISION_CONFIG

    param = param.strip()
    if not param:
        return "[错误] analyze_image_cloud: 参数不能为空，格式为 '图片路径或URL | 提示词'"

    image_path, prompt = _parse_image_tool_args(param)

    _img_exts = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
    _check_ext = os.path.splitext(image_path.split("?")[0])[1].lower()
    if _check_ext and _check_ext not in _img_exts:
        return (
            f"[错误] analyze_image_cloud: 不支持的文件类型 '{_check_ext}'。"
            f"此工具仅接受图片文件 {sorted(_img_exts)}。"
            f"Excel 请用 read_xlsx，PDF 请用 read_pdf。"
        )

    # 选择 provider（GMI 优先，回落 Kimi）
    provider = VISION_CONFIG["provider"]
    prov_cfg = VISION_CONFIG.get(provider, {})
    if not prov_cfg.get("api_key"):
        # 首选 provider 无 key，尝试另一个
        alt = "kimi" if provider == "gmi" else "gmi"
        if VISION_CONFIG.get(alt, {}).get("api_key"):
            provider, prov_cfg = alt, VISION_CONFIG[alt]
        else:
            return (
                "[错误] analyze_image_cloud: 未配置视觉 provider 的 api_key"
                "（需 GMI_API_KEY 或 MOONSHOT_API_KEY）"
            )

    logger.info(
        "analyze_image_cloud 调用 provider=%s path=%s prompt=%s",
        provider, image_path, prompt[:100],
    )

    is_url = image_path.startswith("http://") or image_path.startswith("https://")
    resolved_path = None

    try:
        resolved_path = _resolve_image_path(image_path)
    except (FileNotFoundError, RuntimeError) as e:
        logger.error("analyze_image_cloud: %s", e)
        return f"[错误] analyze_image_cloud: {e}"

    try:
        from openai import OpenAI

        client = OpenAI(
            api_key=prov_cfg["api_key"],
            base_url=prov_cfg["base_url"],
            timeout=120.0,
        )

        with open(resolved_path, "rb") as f:
            image_data = base64.b64encode(f.read()).decode("utf-8")

        ext = resolved_path.rsplit(".", 1)[-1].lower() if "." in resolved_path else "png"
        mime_map = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "webp": "webp", "gif": "gif"}
        mime_type = mime_map.get(ext, "png")
        image_url = f"data:image/{mime_type};base64,{image_data}"

        def _call(model_name: str) -> str:
            resp = client.chat.completions.create(
                model=model_name,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "image_url", "image_url": {"url": image_url}},
                            {"type": "text", "text": prompt},
                        ],
                    },
                ],
                max_tokens=4096,
                stream=False,
            )
            text = resp.choices[0].message.content or ""
            tok = resp.usage.total_tokens if resp.usage else 0
            return text, tok

        t0 = time.time()
        model_used = prov_cfg["model"]
        content, tokens = _call(model_used)

        # GMI 普通模型空响应 → 升级复杂模型重试一次（复杂/模糊图）
        if provider == "gmi" and not content.strip() and prov_cfg.get("model_complex"):
            model_used = prov_cfg["model_complex"]
            logger.info("analyze_image_cloud: flash-lite 空响应，升级 %s 重试", model_used)
            content, tokens = _call(model_used)

        elapsed_ms = (time.time() - t0) * 1000
        content = content or "(视觉模型返回空内容)"

        logger.info(
            "analyze_image_cloud 完成 provider=%s model=%s elapsed=%.0fms tokens=%d",
            provider, model_used, elapsed_ms, tokens,
        )

        return content

    except ImportError:
        logger.error("analyze_image_cloud: openai SDK 未安装")
        return "[错误] analyze_image_cloud: openai SDK 未安装"
    except Exception as exc:
        logger.error("analyze_image_cloud %s 调用失败: %s", provider, exc, exc_info=True)
        return f"[错误] analyze_image_cloud 执行失败: {exc}"
    finally:
        if is_url and resolved_path and os.path.exists(resolved_path):
            os.unlink(resolved_path)


# --------------------------------------------------------------------------
# 网页内容提取（缓存到内存，save_content 负责落盘）
# --------------------------------------------------------------------------

# 会话级累积缓存：每次 tavily_extract 的结果追加到此列表
_session_extracts: List[dict] = []


def get_session_extracts() -> List[dict]:
    """返回本次会话所有 extract 结果（供策展步骤使用）"""
    return _session_extracts


def clear_session_extracts() -> None:
    """清空会话缓存（每次 Agent 运行前调用）"""
    global _session_extracts
    _session_extracts = []


@register(
    name="tavily_extract",
    description="使用 Tavily Extract 提取指定网页的正文内容并缓存到内存；如需持久化保存可调用 save_content 写入本地 Markdown",
    signature="tavily_extract(url)",
    examples=["tavily_extract(https://en.wikipedia.org/wiki/Artificial_intelligence)"],
)
def tavily_extract(url: str) -> str:
    """Tavily 网页内容提取（仅缓存，不自动落盘）

    提取成功后将全文缓存在内存中。LLM 可根据内容价值决定是否调用
    save_content 写入磁盘。Wikipedia 等免费可查内容无需保存，
    行业报告、研究文档、专有数据等才需要。

    Args:
        url: 要提取内容的网页 URL

    Returns:
        标题 + 预览（前 1500 字符）+ 总字数 + 保存提示
    """
    url = url.strip()
    if not url:
        return "[错误] tavily_extract: URL 不能为空"

    logger.info("tavily_extract 调用 url=%s", url)

    tavily_api_key = os.environ.get("TAVILY_API_KEY", "")
    if not tavily_api_key:
        return "[错误] tavily_extract: 未配置 TAVILY_API_KEY，请在 .env 中设置"

    try:
        from tavily import TavilyClient

        client = TavilyClient(api_key=tavily_api_key)
        response = client.extract(url)

        results = response.get("results", [])
        if not results:
            return f"Tavily Extract 未能从 '{url}' 提取到内容。"

        raw_content = results[0].get("raw_content", "")
        title = results[0].get("title", "")
        if not raw_content:
            return f"Tavily Extract: '{url}' 页面无正文内容。"

        # 追加到会话级缓存，不落盘
        _session_extracts.append({
            "url": url, "title": title, "raw_content": raw_content,
        })

        preview = raw_content[:1500]
        if len(raw_content) > 1500:
            preview += "..."

        logger.info("tavily_extract 完成 url=%s len=%d cached=true", url, len(raw_content))
        return (
            f"Tavily Extract - {title or url}\n"
            f"总字符数: {len(raw_content)}\n\n"
            f"--- 内容预览（前 1500 字符）---\n{preview}\n\n"
            f"💡 全文已缓存。如果内容有长期参考价值（如行业报告、研究文档），可调用 save_content(文件名) 保存到本地。"
        )

    except ImportError:
        logger.error("tavily_extract: tavily-python 未安装")
        return "[错误] tavily_extract: tavily-python 未安装，请执行 pip install tavily-python"
    except Exception as exc:
        logger.error("tavily_extract 失败: %s", exc, exc_info=True)
        return f"[错误] tavily_extract 执行失败: {exc}"


# --------------------------------------------------------------------------
# 网页正文提取 — Jina Reader + trafilatura 双引擎
# --------------------------------------------------------------------------

def extract_url_text(url: str, timeout: int = 30, use_fallback: bool = True) -> tuple[str, str]:
    """从 URL 提取正文纯文本 — Jina Reader 优先，trafilatura 兜底。

    这是 web_fetch 的核心提取逻辑，抽出供搜索增强层等复用。
    调用方需自行保证 url 已规范化（带 scheme）。

    Args:
        url: 目标 URL
        timeout: Jina Reader 请求超时（秒）
        use_fallback: Jina 失败时是否回退 trafilatura（增强层可关掉以提速）

    Returns:
        (正文, 引擎名)；均失败时返回 ("", "")
    """
    # ── 引擎 1: Jina Reader ──
    try:
        resp = requests.get(
            f"https://r.jina.ai/{url}",
            headers={"Accept": "text/markdown", "User-Agent": "NexaAgent-ReAct/1.0"},
            timeout=timeout,
        )
        if resp.status_code == 200 and resp.text.strip():
            return resp.text.strip(), "Jina Reader"
        logger.info("extract_url_text Jina 返回空/失败 status=%d", resp.status_code)
    except Exception as exc:
        logger.info("extract_url_text Jina 异常: %s", exc)

    # ── 引擎 2: trafilatura ──
    if use_fallback:
        try:
            import trafilatura

            downloaded = trafilatura.fetch_url(url)
            if downloaded:
                extracted = trafilatura.extract(
                    downloaded,
                    include_links=False,
                    include_images=False,
                    include_tables=True,
                    output_format="markdown",
                )
                if extracted and extracted.strip():
                    return extracted.strip(), "trafilatura"
        except ImportError:
            logger.warning("extract_url_text: trafilatura 未安装")
        except Exception as exc:
            logger.error("extract_url_text trafilatura 失败: %s", exc)

    return "", ""


@register(
    name="web_fetch",
    description="提取指定网页的正文内容。优先使用 Jina Reader（免费），失败时自动回退 trafilatura 本地解析。适合阅读文章、文档等深层内容，不依赖 Tavily API Key。",
    signature="web_fetch(url)",
    examples=["web_fetch(https://en.wikipedia.org/wiki/Artificial_intelligence)"],
)
def web_fetch(url: str) -> str:
    """网页正文提取 — Jina Reader 优先，trafilatura 兜底

    不依赖 Tavily API Key，可免费使用。
    提取正文后自动追加到会话缓存，可配合 save_content 持久化。

    Args:
        url: 要提取内容的网页 URL

    Returns:
        标题 + 正文预览（前 2000 字符）+ 总字符数
    """
    url = url.strip()
    if not url:
        return "[错误] web_fetch: URL 不能为空"

    # 拒绝包含空格的输入（通常是误把命令当 URL）
    if " " in url:
        return f"[错误] web_fetch: URL 含空格，疑似非 URL 输入: '{url[:80]}'。请提供完整的 http/https URL。"

    # 确保 URL 有 scheme
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    # 验证 netloc 存在（拒绝 file:// 等混合 scheme）
    from urllib.parse import urlparse as _urlparse
    _p = _urlparse(url)
    if not _p.netloc or _p.netloc.startswith("file:"):
        return f"[错误] web_fetch: 无效 URL '{url[:100]}'。请提供标准的 http/https URL（如 https://example.com/page）。"

    logger.info("web_fetch 调用 url=%s", url)

    content, engine = extract_url_text(url, timeout=30)
    if content:
        logger.info("web_fetch 成功 engine=%s len=%d", engine, len(content))

    if not content:
        return (
            f"[错误] web_fetch: 无法从 '{url}' 提取内容。\n"
            f"Jina Reader 和 trafilatura 均未返回有效正文。"
        )

    # 限制长度
    max_len = 8000
    if len(content) > max_len:
        content_preview = content[:max_len]
    else:
        content_preview = content

    # 追加到会话缓存
    _session_extracts.append({
        "url": url, "title": url, "raw_content": content,
    })

    return (
        f"web_fetch ({engine}) — {url}\n"
        f"总字符数: {len(content)}\n\n"
        f"--- 正文预览（前 {min(len(content), max_len)} 字符）---\n"
        f"{content_preview}"
        f"{'...(截断)' if len(content) > max_len else ''}\n\n"
        f"💡 全文已缓存。如需保存可调用 save_content(文件名)。"
    )


# --------------------------------------------------------------------------
# 内容保存（从缓存写入 .md）
# --------------------------------------------------------------------------

def write_extract_to_disk(extract: dict, filename: str) -> str:
    """将一条 extract 写入 data/extracts/{filename}.md

    供 react_agent 策展步骤直接调用，也供 save_content 工具复用。

    Args:
        extract: {"url", "title", "raw_content"}
        filename: 文件名（不含路径和 .md）

    Returns:
        保存的文件路径
    """
    safe_name = re.sub(r"[^\w\-一-鿿]", "_", filename).strip("_")[:80]
    if not safe_name:
        raise ValueError(f"文件名无效: {filename}")

    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    extract_dir = os.path.join(project_root, "data", "extracts")
    os.makedirs(extract_dir, exist_ok=True)

    filepath = os.path.join(extract_dir, f"{safe_name}.md")

    extracted_at = datetime.now().isoformat()
    md_content = (
        f"---\n"
        f"title: {extract.get('title') or 'Extracted Content'}\n"
        f"source: {extract.get('url', '')}\n"
        f"extracted_at: {extracted_at}\n"
        f"content_length: {len(extract.get('raw_content', ''))}\n"
        f"---\n\n"
        f"{extract.get('raw_content', '')}"
    )

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(md_content)

    logger.info("write_extract_to_disk file=%s len=%d", filepath, len(extract.get("raw_content", "")))
    return filepath


@register(
    name="save_content",
    description="将最近一次 tavily_extract 缓存的网页全文写入本地 Markdown 文件（含 YAML frontmatter）。注意：Wikipedia 等免费可查内容无需保存，仅行业报告、研究文档、公司数据等有长期参考价值的内容才需要持久化。",
    signature="save_content(filename)",
    examples=["save_content(apple_ghg_emissions_2025)"],
)
def save_content(filename: str) -> str:
    """将最近一条 extract 写入 .md 文件

    供 ReAct 循环中的 LLM 直接调用。策展步骤（_curation_step）中则使用
    write_extract_to_disk() 逐条写入。

    Args:
        filename: 文件名（不含路径和 .md），如 "apple_ghg_2025"

    Returns:
        保存路径及元信息
    """
    filename = filename.strip()
    if not filename:
        return "[错误] save_content: 文件名不能为空"

    if not _session_extracts:
        return (
            "[错误] save_content: 没有可保存的内容。"
            "请先调用 tavily_extract(url) 提取网页内容后再保存。"
        )

    # 取最近一条 extract
    extract = _session_extracts[-1]

    try:
        filepath = write_extract_to_disk(extract, filename)
        title = extract.get("title", "")
        raw_content = extract.get("raw_content", "")
        return (
            f"✅ 已保存到: {filepath}\n"
            f"标题: {title}\n"
            f"字符数: {len(raw_content)}"
        )

    except Exception as exc:
        logger.error("save_content 失败: %s", exc, exc_info=True)
        return f"[错误] save_content 写入失败: {exc}"


# --------------------------------------------------------------------------
# 数学计算
# --------------------------------------------------------------------------

@register(
    name="calculator",
    description="安全地计算数学表达式，支持 sqrt/log/sin/cos/pow 等函数",
    signature="calculator(expression)",
    examples=["calculator(sqrt(144) + pow(2, 10))"],
)
def calculator(expression: str) -> str:
    """数学表达式计算器

    使用 numexpr 进行安全计算，支持常见数学函数。

    Args:
        expression: 数学表达式字符串

    Returns:
        计算结果
    """
    import math as _math

    expression = expression.strip()
    if not expression:
        return "[错误] calculator: 表达式不能为空"

    # 预处理: LLM 常用 ^ 表示幂运算，Python 里需要 **
    expression = expression.replace("^", "**")
    # 预处理: 去掉多语句（numexpr 不支持分号分隔）
    if ";" in expression:
        parts = [p.strip() for p in expression.split(";") if p.strip()]
        expression = parts[-1] if parts else expression

    logger.info("calculator 调用 expression=%s", expression)

    # 安全的数学函数和常量（eval 和 numexpr 共用）
    _safe_dict = {
        "sqrt": _math.sqrt,
        "log": _math.log,
        "log10": _math.log10,
        "log2": _math.log2,
        "sin": _math.sin,
        "cos": _math.cos,
        "tan": _math.tan,
        "asin": _math.asin,
        "acos": _math.acos,
        "atan": _math.atan,
        "atan2": _math.atan2,
        "pow": pow,
        "abs": abs,
        "pi": _math.pi,
        "e": _math.e,
    }

    try:
        import numexpr

        result = numexpr.evaluate(expression, local_dict=_safe_dict)
        if hasattr(result, "item"):
            result = result.item()

        logger.info("calculator 结果=%s", result)
        return f"计算结果: {expression} = {result}"

    except ImportError:
        pass
    except Exception as exc:
        logger.debug("calculator numexpr 失败，降级到 eval: %s", exc)

    # eval fallback
    try:
        result = eval(expression, {"__builtins__": {}}, _safe_dict)
        logger.info("calculator 结果=%s", result)
        return f"计算结果: {expression} = {result}"
    except Exception as e2:
        logger.error("calculator 计算失败: %s", e2)
        return f"[错误] calculator: 计算失败 - {e2}"


# --------------------------------------------------------------------------
# 时间查询
# --------------------------------------------------------------------------

@register(
    name="get_current_time",
    description="返回当前本地日期和时间",
    signature="get_current_time()",
    examples=["get_current_time()"],
)
def get_current_time(_param: str = "") -> str:
    """获取当前本地时间

    Returns:
        格式化的日期时间字符串
    """
    logger.info("get_current_time 调用")
    now = datetime.now()
    weekday_map = {
        0: "星期一", 1: "星期二", 2: "星期三",
        3: "星期四", 4: "星期五", 5: "星期六", 6: "星期日",
    }
    weekday = weekday_map[now.weekday()]
    result = now.strftime(f"当前时间: %Y年%m月%d日 {weekday} %H:%M:%S")
    return result


# --------------------------------------------------------------------------
# 域名核验（OfferCheck 领域工具）
# --------------------------------------------------------------------------

def _extract_domain(raw: str) -> str:
    """从 URL / 邮箱 / 裸域名中提取注册域名（eTLD+1 的近似）"""
    s = raw.strip().lower()
    # 邮箱 → 取 @ 之后
    if "@" in s:
        s = s.split("@", 1)[1]
    # URL → 取 host
    s = re.sub(r"^[a-z]+://", "", s)
    s = s.split("/", 1)[0].split("?", 1)[0]
    s = s.split(":", 1)[0]  # 去端口
    # 去掉常见子域前缀，保留主域（简化：取最后两段；多级公共后缀交给 RDAP 兜底）
    parts = [p for p in s.split(".") if p]
    if len(parts) >= 3 and parts[0] in ("www", "mail", "smtp", "careers", "jobs", "hr", "apply"):
        parts = parts[1:]
    if len(parts) >= 2:
        return ".".join(parts[-2:])
    return s


@register(
    name="domain_whois_lookup",
    description=(
        "查询域名的注册信息（注册时间、注册商、有效期），用于识别仿冒真实企业的山寨域名。"
        "近期新注册的域名冒用知名公司名是招聘诈骗的典型信号。基于免费的 RDAP 协议，无需 API Key。"
    ),
    signature="domain_whois_lookup(domain_or_url_or_email)",
    examples=[
        "domain_whois_lookup(example.com)",
        "domain_whois_lookup(https://careers.acme-corp.io/apply)",
        "domain_whois_lookup(hr@suspicious-google-jobs.com)",
    ],
)
def domain_whois_lookup(param: str) -> str:
    """通过 RDAP 查询域名注册信息，辅助识别山寨/仿冒域名。

    Args:
        param: 域名，或包含域名的 URL / 邮箱地址

    Returns:
        注册商、注册/更新/到期时间、域名状态，以及基于注册时长的风险提示
    """
    raw = (param or "").strip()
    if not raw:
        return "[错误] domain_whois_lookup: 域名不能为空"

    domain = _extract_domain(raw)
    logger.info("domain_whois_lookup 调用 raw=%s domain=%s", raw[:80], domain)

    # RDAP：先走 IANA bootstrap 的通用入口，rdap.org 会按 TLD 重定向到权威 RDAP 服务
    url = f"https://rdap.org/domain/{domain}"
    try:
        resp = requests.get(
            url,
            timeout=15,
            headers={"Accept": "application/rdap+json", "User-Agent": "nexa-agent-offercheck/1.0"},
            allow_redirects=True,
        )
    except Exception as exc:
        logger.error("domain_whois_lookup 请求失败: %s", exc)
        return f"[错误] domain_whois_lookup 查询 {domain} 失败: {exc}"

    if resp.status_code == 404:
        return (
            f"域名 {domain} 在 RDAP 中未找到注册记录。"
            f"可能是：1) 该 TLD 不支持 RDAP；2) 域名未注册（⚠️ 高度可疑，冒用邮箱常用未注册域名）；"
            f"3) 域名拼写有误。建议结合 web_search 进一步核实该域名是否为公司官方域名。"
        )
    if resp.status_code != 200:
        return f"[错误] domain_whois_lookup: RDAP 返回状态码 {resp.status_code}（域名 {domain}）"

    try:
        data = resp.json()
    except Exception as exc:
        return f"[错误] domain_whois_lookup: 无法解析 RDAP 响应: {exc}"

    # 解析事件（注册/到期/更新时间）
    events = {e.get("eventAction"): e.get("eventDate") for e in data.get("events", []) if isinstance(e, dict)}
    registration = events.get("registration")
    expiration = events.get("expiration")
    last_changed = events.get("last changed") or events.get("last update of RDAP database")

    # 注册商（entities 中 role 含 registrar）
    registrar = "未知"
    for ent in data.get("entities", []) or []:
        roles = ent.get("roles", []) if isinstance(ent, dict) else []
        if "registrar" in roles:
            vcard = ent.get("vcardArray")
            if isinstance(vcard, list) and len(vcard) > 1:
                for item in vcard[1]:
                    if isinstance(item, list) and len(item) >= 4 and item[0] == "fn":
                        registrar = item[3]
                        break
            if registrar == "未知":
                registrar = ent.get("handle", "未知")
            break

    statuses = ", ".join(data.get("status", [])) or "未标注"

    # 基于注册时长的风险提示
    risk_note = ""
    if registration:
        try:
            reg_dt = datetime.fromisoformat(registration.replace("Z", "+00:00"))
            age_days = (datetime.now(reg_dt.tzinfo) - reg_dt).days
            years = age_days / 365.25
            if age_days < 90:
                risk_note = (
                    f"\n⚠️ 高风险：该域名注册于 {age_days} 天前（不足 3 个月）。"
                    f"若它自称是知名/成熟企业的官方域名，这是仿冒山寨域名的强烈信号。"
                )
            elif age_days < 365:
                risk_note = f"\n⚠️ 注意：该域名注册不足 1 年（约 {age_days} 天），对成熟企业而言偏新，建议交叉核实。"
            else:
                risk_note = f"\n✅ 该域名已注册约 {years:.1f} 年，注册时长本身不构成红旗（仍需结合其他证据）。"
        except Exception:
            pass

    lines = [
        f"域名 {domain} 的 RDAP 注册信息：",
        f"- 注册商: {registrar}",
        f"- 注册时间: {registration or '未提供'}",
        f"- 到期时间: {expiration or '未提供'}",
        f"- 最近变更: {last_changed or '未提供'}",
        f"- 域名状态: {statuses}",
    ]
    result = "\n".join(lines) + risk_note
    logger.info("domain_whois_lookup 完成 domain=%s registrar=%s reg=%s", domain, registrar, registration)
    return result


# --------------------------------------------------------------------------
# PDF 解析
# --------------------------------------------------------------------------

@register(
    name="read_pdf",
    description="解析 PDF 文件并提取正文为 Markdown 格式。支持本地路径和 URL（自动下载）。"
                "适合政府文档、学术论文、报告等数字 PDF。"
                "对于大文档，可指定页码范围：read_pdf(path_or_url | pages=3-5)",
    signature="read_pdf(path_or_url | pages=start-end)",
    examples=[
        "read_pdf(https://example.com/report.pdf)",
        "read_pdf(data/document.pdf | pages=1-5)",
    ],
)
def read_pdf(param: str) -> str:
    """解析 PDF 并提取 Markdown 文本

    支持:
    - 本地文件路径或 URL
    - 可选分页: path_or_url | pages=3-5

    策略：小文档全量返回，大文档返回目录概要+指引 Agent 分页精读。
    """
    param = param.strip()
    if not param:
        return "[错误] read_pdf: 参数不能为空，请提供 PDF 路径或 URL"

    # 解析可选的 pages 参数
    page_start, page_end = None, None
    source = param
    if "|" in param:
        parts = param.split("|", 1)
        source = parts[0].strip()
        extra = parts[1].strip()
        import re as _re
        pages_match = _re.search(r"pages?\s*=\s*(\d+)(?:\s*-\s*(\d+))?", extra, _re.IGNORECASE)
        if pages_match:
            page_start = int(pages_match.group(1))
            page_end = int(pages_match.group(2)) if pages_match.group(2) else page_start

    logger.info("read_pdf 调用 source=%s pages=%s-%s", source[:150], page_start, page_end)

    import tempfile

    tmp_path = None
    pdf_path = None

    try:
        if source.startswith("http://") or source.startswith("https://"):
            headers = {
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            }
            resp = requests.get(source, timeout=60, headers=headers)
            if resp.status_code != 200:
                return f"[错误] read_pdf: 下载失败 HTTP {resp.status_code} — {source[:100]}"
            if len(resp.content) < 100:
                return f"[错误] read_pdf: 下载内容太小 ({len(resp.content)} bytes)"

            tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            tmp.write(resp.content)
            tmp.close()
            tmp_path = tmp.name
            pdf_path = tmp_path
            logger.info("read_pdf 已下载 %d bytes → %s", len(resp.content), tmp_path)
        else:
            _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            for c in [source, os.path.join(_root, source), os.path.join(_root, "data", source)]:
                if os.path.isfile(c):
                    pdf_path = c
                    break
            if not pdf_path:
                return f"[错误] read_pdf: 文件不存在 — {source}"

        import pymupdf4llm
        import pymupdf

        doc = pymupdf.open(pdf_path)
        total_pages = len(doc)
        doc.close()

        # 分页提取
        kwargs = {}
        if page_start is not None:
            pages_list = list(range(page_start - 1, min(page_end or page_start, total_pages)))
            kwargs["pages"] = pages_list

        md_text = pymupdf4llm.to_markdown(pdf_path, **kwargs)

        if not md_text or len(md_text.strip()) < 20:
            return f"[错误] read_pdf: PDF 提取结果为空或过短，可能是扫描件或加密文件"

        logger.info("read_pdf 提取完成 total_pages=%d chars=%d", total_pages, len(md_text))

        _session_extracts.append({
            "url": source,
            "title": os.path.basename(source)[:80],
            "raw_content": md_text,
        })

        page_info = f"第 {page_start}-{page_end} 页" if page_start else "全文"
        return f"PDF 解析成功 (共 {total_pages} 页, {page_info}, {len(md_text)} 字符):\n\n{md_text}"

    except Exception as exc:
        logger.error("read_pdf 失败: %s", exc, exc_info=True)
        return f"[错误] read_pdf: 解析失败 — {exc}"
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


# --------------------------------------------------------------------------
# Excel (xlsx/xls/csv) 解析
# --------------------------------------------------------------------------

@register(
    name="read_xlsx",
    description="解析 Excel (.xlsx/.xls) 或 CSV 文件，将内容转为 Markdown 表格。"
                "支持本地路径和 URL（自动下载）。多 Sheet 自动分节展示。"
                "大文件默认显示前 100 行，可指定行范围：read_xlsx(path | rows=10-50)",
    signature="read_xlsx(path_or_url | rows=start-end)",
    examples=[
        "read_xlsx(data/stats.xlsx)",
        "read_xlsx(https://example.com/data.xlsx | rows=1-20)",
    ],
)
def read_xlsx(param: str) -> str:
    param = param.strip()
    if not param:
        return "[错误] read_xlsx: 参数不能为空，请提供文件路径或 URL"

    row_start, row_end = None, None
    source = param
    if "|" in param:
        parts = param.split("|", 1)
        source = parts[0].strip()
        extra = parts[1].strip()
        rows_match = re.search(r"rows?\s*=\s*(\d+)(?:\s*-\s*(\d+))?", extra, re.IGNORECASE)
        if rows_match:
            row_start = int(rows_match.group(1))
            row_end = int(rows_match.group(2)) if rows_match.group(2) else row_start

    logger.info("read_xlsx 调用 source=%s rows=%s-%s", source[:150], row_start, row_end)

    import tempfile

    tmp_path = None
    file_path = None

    try:
        if source.startswith("http://") or source.startswith("https://"):
            headers = {
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            }
            resp = requests.get(source, timeout=60, headers=headers)
            if resp.status_code != 200:
                return f"[错误] read_xlsx: 下载失败 HTTP {resp.status_code} — {source[:100]}"

            suffix = ".xlsx"
            if source.lower().endswith(".csv"):
                suffix = ".csv"
            elif source.lower().endswith(".xls"):
                suffix = ".xls"
            tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
            tmp.write(resp.content)
            tmp.close()
            tmp_path = tmp.name
            file_path = tmp_path
            logger.info("read_xlsx 已下载 %d bytes → %s", len(resp.content), tmp_path)
        else:
            _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            for c in [source, os.path.join(_root, source), os.path.join(_root, "data", source)]:
                if os.path.isfile(c):
                    file_path = c
                    break
            if not file_path:
                return f"[错误] read_xlsx: 文件不存在 — {source}"

        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".csv":
            return _parse_csv(file_path, row_start, row_end)
        else:
            return _parse_excel(file_path, row_start, row_end)

    except Exception as exc:
        logger.error("read_xlsx 失败: %s", exc, exc_info=True)
        return f"[错误] read_xlsx: 解析失败 — {exc}"
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _parse_excel(file_path: str, row_start: Optional[int], row_end: Optional[int]) -> str:
    try:
        import openpyxl
    except ImportError:
        return "[错误] read_xlsx: 需要安装 openpyxl (pip install openpyxl)"

    from collections import defaultdict

    # read_only=False required to access cell fill/color attributes
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    sections = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # values_only=False to access cell styling (fill colors)
        all_cell_rows = list(ws.iter_rows(values_only=False))
        rows = [tuple(cell.value for cell in row) for row in all_cell_rows]
        total_rows += len(rows)

        if not rows:
            sections.append(f"### Sheet: {sheet_name}\n(空表)")
            continue

        # 行范围过滤（1-indexed，第 1 行视为表头）
        if row_start is not None:
            start_idx = max(0, row_start - 1)
            end_idx = min(len(rows), (row_end or row_start))
            header_vals = rows[0] if start_idx > 0 else None
            header_cells = all_cell_rows[0] if start_idx > 0 else None
            selected_vals = rows[start_idx:end_idx]
            selected_cells = all_cell_rows[start_idx:end_idx]
            if header_vals and selected_vals and selected_vals[0] != header_vals:
                selected_vals = [header_vals] + selected_vals
                selected_cells = [header_cells] + selected_cells
            truncated = False
        else:
            max_rows = 100
            selected_vals = rows[:max_rows]
            selected_cells = all_cell_rows[:max_rows]
            truncated = len(rows) > max_rows

        md_lines = _rows_to_markdown(selected_vals)
        section = f"### Sheet: {sheet_name} ({len(rows)} 行)\n\n" + "\n".join(md_lines)

        # 提取单元格背景色，按颜色分组列出坐标
        color_map: dict = defaultdict(list)
        for cell_row in selected_cells:
            for cell in cell_row:
                try:
                    fill = cell.fill
                    if fill.patternType and fill.patternType not in ("none", None):
                        fg = fill.fgColor
                        if fg.type == "rgb":
                            rgb6 = fg.rgb[-6:].upper()
                            if rgb6 not in ("FFFFFF", "000000"):
                                color_map[f"#{rgb6}"].append(
                                    f"{cell.column_letter}{cell.row}"
                                )
                        elif fg.type == "theme":
                            color_map[f"theme:{fg.theme}"].append(
                                f"{cell.column_letter}{cell.row}"
                            )
                except Exception:
                    pass

        if color_map:
            color_lines = []
            for color, coords in sorted(color_map.items()):
                coords_str = ", ".join(coords[:30])
                if len(coords) > 30:
                    coords_str += f" ... (共 {len(coords)} 个)"
                color_lines.append(f"  {color}: {coords_str}")
            section += "\n\n**单元格背景色分布:**\n" + "\n".join(color_lines)

        if truncated:
            section += f"\n\n... 仅显示前 100 行，共 {len(rows)} 行。使用 `rows=101-200` 查看更多。"
        sections.append(section)

    wb.close()
    result = "\n\n".join(sections)
    logger.info(
        "read_xlsx 解析完成 sheets=%d total_rows=%d chars=%d",
        len(wb.sheetnames), total_rows, len(result),
    )
    return f"Excel 解析成功 ({len(wb.sheetnames)} 个 Sheet, {total_rows} 行):\n\n{result}"


def _parse_csv(file_path: str, row_start: Optional[int], row_end: Optional[int]) -> str:
    import csv as _csv
    rows = []
    with open(file_path, encoding="utf-8", errors="replace", newline="") as f:
        reader = _csv.reader(f)
        for row in reader:
            rows.append(tuple(row))

    if not rows:
        return "CSV 文件为空"

    if row_start is not None:
        start_idx = max(0, row_start - 1)
        end_idx = min(len(rows), (row_end or row_start))
        header = rows[0] if start_idx > 0 else None
        selected = rows[start_idx:end_idx]
        if header and selected and selected[0] != header:
            selected = [header] + selected
        truncated = False
    else:
        selected = rows[:100]
        truncated = len(rows) > 100

    md_lines = _rows_to_markdown(selected)
    result = "\n".join(md_lines)
    if truncated:
        result += f"\n\n... 仅显示前 100 行，共 {len(rows)} 行。使用 `rows=101-200` 查看更多。"

    return f"CSV 解析成功 ({len(rows)} 行):\n\n{result}"


def _rows_to_markdown(rows: list) -> List[str]:
    if not rows:
        return []

    def cell_str(v):
        if v is None:
            return ""
        return str(v).replace("|", "\\|").replace("\n", " ")

    header = rows[0]
    n_cols = len(header)
    lines = [
        "| " + " | ".join(cell_str(c) for c in header) + " |",
        "| " + " | ".join("---" for _ in range(n_cols)) + " |",
    ]
    for row in rows[1:]:
        padded = list(row) + [None] * (n_cols - len(row)) if len(row) < n_cols else row[:n_cols]
        lines.append("| " + " | ".join(cell_str(c) for c in padded) + " |")
    return lines


# ==========================================================================
# 工具执行入口
# ==========================================================================

def execute_tool(tool_name: str, tool_args: str) -> str:
    """根据工具名和参数执行对应的工具函数

    Args:
        tool_name: 工具名称（必须已在 TOOLS 注册表中）
        tool_args: 参数字符串（原样传给工具函数）

    Returns:
        工具执行结果字符串（Observation）
    """
    if tool_name not in TOOLS:
        available = ", ".join(TOOLS.keys())
        return f"[错误] 未知工具: {tool_name}。可用工具: {available}"

    tool_func = TOOLS[tool_name]
    try:
        result = tool_func(tool_args)
        return result
    except Exception as exc:
        logger.error(
            "工具 %s 执行异常: %s", tool_name, exc, exc_info=True
        )
        return f"[错误] 工具 {tool_name} 执行失败: {exc}"


def get_tools_description() -> str:
    """生成工具列表的描述文本（用于 System Prompt）"""
    lines = []
    for name, meta in TOOL_META.items():
        lines.append(f"- **{meta['signature']}**: {meta['description']}")
        for ex in meta["examples"]:
            lines.append(f"  - 示例: `{ex}`")
    return "\n".join(lines)
